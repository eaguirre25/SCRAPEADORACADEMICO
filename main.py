#!/usr/bin/env python3
"""
Scraper acumulativo sobre dirección escolar / gestión escolar.

Fuentes:
- OpenAlex (API REST)
- CONICET Digital (OAI-PMH)
- SEDICI - UNLP (OAI-PMH)
- RIAA - UNSAM (OAI-PMH)

Deduplicación por capas:
1. DOI exacto
2. Similitud de título > 92% + mismo año
3. Similitud de título > 80% + primer autor similar + mismo año
"""

from __future__ import annotations

import csv
import json
import os
import re
import smtplib
import time
import xml.etree.ElementTree as ET
from datetime import date
from difflib import SequenceMatcher
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import requests
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from drive_uploader import DriveUploader
    DRIVE_AVAILABLE = True
except ImportError:
    DRIVE_AVAILABLE = False

# ── Configuración ─────────────────────────────────────────────────────────────

SEARCH_TERMS: List[str] = [
    "gestión escolar",
    "dirección escolar",
    "gestión educativa",
    "school management",
    "educational leadership",
]

OAI_SOURCES = [
    {
        "name":     "SEDICI-UNLP",
        "url":      "https://sedici.unlp.edu.ar/oai/request",
        "max_pages": 60,
    },
    {
        "name":     "RIAA-UNSAM",
        "url":      "https://repositorio.unsam.edu.ar/oai/request",
        "max_pages": 30,
    },
]

# Términos ampliados para OAI-PMH
OAI_SEARCH_TERMS: List[str] = [
    "gestión escolar", "dirección escolar", "gestión educativa",
    "school management", "educational leadership",
    "school principal", "principalship", "headteacher",
    "conducción escolar", "administración escolar",
    "director escolar", "liderazgo educativo",
    "liderazgo escolar", "director de escuela",
]

# CONICET Digital: usa su API REST de DSpace en lugar de OAI-PMH
CONICET_REST_URL = "https://ri.conicet.gov.ar/rest"
CONICET_SEARCH_TERMS: List[str] = [
    "gestion escolar", "direccion escolar", "gestion educativa",
    "school management", "educational leadership",
    "school principal", "director escolar", "liderazgo educativo",
    "liderazgo escolar", "conduccion escolar",
]

START_DATE         = "2020-01-01"
MAX_PAGES_PER_TERM: Optional[int] = 3

# Umbrales de deduplicación fuzzy
FUZZY_TITLE_HIGH   = 0.92   # título solo → duplicado
FUZZY_TITLE_LOW    = 0.80   # título + autor → duplicado
FUZZY_AUTHOR_MIN   = 0.85

EMAIL_ITEM_LIMIT   = 50

BASE_DIR   = Path(__file__).resolve().parent
DATA_DIR   = BASE_DIR / "data"
MASTER_CSV = DATA_DIR / "master_records.csv"
SEEN_IDS_JSON = DATA_DIR / "seen_ids.json"
EXCEL_FILE = DATA_DIR / "publicaciones.xlsx"
PDFS_DIR   = DATA_DIR / "pdfs"

CSV_FIELDS = [
    "record_id", "first_seen_date", "search_term", "source",
    "origin", "document_type", "authors", "title", "abstract",
    "keywords", "publication_year", "publication_date",
    "doi", "url", "openalex_id", "is_oa", "pdf_url",
]

# Namespaces OAI-PMH
OAI_NS = {
    "oai":    "http://www.openarchives.org/OAI/2.0/",
    "dc":     "http://purl.org/dc/elements/1.1/",
    "oai_dc": "http://www.openarchives.org/OAI/2.0/oai_dc/",
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def as_dict(value: Any) -> Dict[str, Any]:
    return value if isinstance(value, dict) else {}


def normalize_text(text: Optional[str]) -> str:
    if not text:
        return ""
    text = text.casefold().strip()
    text = re.sub(r"\s+", " ", text)
    # quitar puntuación para comparación
    text = re.sub(r"[^\w\s]", "", text)
    return text


def normalize_doi(doi_value: Optional[str]) -> str:
    if not doi_value:
        return ""
    doi_value = doi_value.strip()
    if "doi.org/" in doi_value:
        doi_value = doi_value.split("doi.org/")[-1]
    return doi_value.lower().strip()


def build_signature(title: Optional[str], year: Any) -> str:
    return f"{normalize_text(title)}::{str(year or '').strip()}"


def get_first_author(authors_str: str) -> str:
    if not authors_str:
        return ""
    return normalize_text(authors_str.split(";")[0].strip())


def build_headers() -> Dict[str, str]:
    gmail_user = os.getenv("GMAIL_USER", "").strip()
    email = gmail_user or "academic-scraper@example.com"
    return {"User-Agent": f"academic-scraper/3.0 (mailto:{email})"}


# ── Persistencia ──────────────────────────────────────────────────────────────

def ensure_storage() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    PDFS_DIR.mkdir(parents=True, exist_ok=True)
    if not MASTER_CSV.exists():
        with MASTER_CSV.open("w", encoding="utf-8", newline="") as f:
            csv.DictWriter(f, fieldnames=CSV_FIELDS).writeheader()
    if not SEEN_IDS_JSON.exists():
        SEEN_IDS_JSON.write_text("{}", encoding="utf-8")


def load_seen_ids() -> Set[str]:
    try:
        data = json.loads(SEEN_IDS_JSON.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            return set(data.keys())
    except Exception:
        pass
    return set()


def save_seen_ids(seen_ids: Set[str]) -> None:
    SEEN_IDS_JSON.write_text(
        json.dumps({k: True for k in sorted(seen_ids)}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def load_existing_signatures() -> Set[str]:
    sigs: Set[str] = set()
    if not MASTER_CSV.exists():
        return sigs
    try:
        with MASTER_CSV.open("r", encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                sigs.add(build_signature(row.get("title", ""), row.get("publication_year", "")))
    except Exception as e:
        print(f"Error leyendo signatures: {e}")
    return sigs


def build_fuzzy_index() -> Dict[str, List[Tuple[str, str]]]:
    """Índice año → [(título_norm, primer_autor)] para dedup fuzzy."""
    index: Dict[str, List[Tuple[str, str]]] = {}
    if not MASTER_CSV.exists():
        return index
    try:
        with MASTER_CSV.open("r", encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                year = str(row.get("publication_year", "") or "").strip()
                t    = normalize_text(row.get("title", ""))
                a    = get_first_author(row.get("authors", ""))
                index.setdefault(year, []).append((t, a))
    except Exception as e:
        print(f"Error construyendo índice fuzzy: {e}")
    return index


def is_fuzzy_duplicate(
    record: Dict[str, Any],
    fuzzy_index: Dict[str, List[Tuple[str, str]]],
) -> bool:
    """
    Tres capas de dedup:
    1. DOI exacto → manejado en seen_ids (antes de llamar esta función)
    2. sim(título) > 0.92 + mismo año → duplicado
    3. sim(título) > 0.80 + sim(primer autor) > 0.85 + mismo año → duplicado
    """
    year       = str(record.get("publication_year", "") or "").strip()
    title_norm = normalize_text(record.get("title", ""))
    first_auth = get_first_author(record.get("authors", ""))

    if not title_norm or not year:
        return False

    for (ex_title, ex_author) in fuzzy_index.get(year, []):
        if not ex_title:
            continue
        sim = SequenceMatcher(None, title_norm, ex_title).ratio()
        if sim >= FUZZY_TITLE_HIGH:
            return True
        if sim >= FUZZY_TITLE_LOW and first_auth and ex_author:
            if SequenceMatcher(None, first_auth, ex_author).ratio() >= FUZZY_AUTHOR_MIN:
                return True
    return False


def append_master_records(records: List[Dict[str, Any]]) -> None:
    if not records:
        return
    # Detectar campos reales del CSV (compatibilidad con versiones anteriores)
    existing_fields = CSV_FIELDS
    if MASTER_CSV.exists():
        try:
            with MASTER_CSV.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                if reader.fieldnames:
                    existing_fields = list(reader.fieldnames)
        except Exception:
            pass
    with MASTER_CSV.open("a", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=existing_fields, extrasaction="ignore")
        for record in records:
            writer.writerow({k: record.get(k, "") for k in existing_fields})


def count_master_records() -> int:
    if not MASTER_CSV.exists():
        return 0
    try:
        with MASTER_CSV.open("r", encoding="utf-8-sig", newline="") as f:
            return sum(1 for _ in csv.DictReader(f))
    except Exception:
        return 0


def read_all_records() -> List[Dict[str, str]]:
    if not MASTER_CSV.exists():
        print("CSV maestro no encontrado.")
        return []
    try:
        csv.field_size_limit(10_000_000)
        with MASTER_CSV.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            print(f"CSV leído: {len(rows)} registros")
            return rows
    except Exception as e:
        print(f"Error leyendo CSV maestro: {e}")
        return []


# ── OpenAlex ──────────────────────────────────────────────────────────────────

def reconstruct_abstract(abstract_inverted_index: Any) -> str:
    if not isinstance(abstract_inverted_index, dict):
        return ""
    tokens: List[Tuple[int, str]] = []
    for word, positions in abstract_inverted_index.items():
        if not isinstance(positions, list):
            continue
        for pos in positions:
            if isinstance(pos, int):
                tokens.append((pos, word))
    if not tokens:
        return ""
    tokens.sort(key=lambda x: x[0])
    return " ".join(w for _, w in tokens)


def extract_authors(work: Dict[str, Any]) -> str:
    authorships = work.get("authorships", [])
    if not isinstance(authorships, list):
        return ""
    names: List[str] = []
    for authorship in authorships:
        if not isinstance(authorship, dict):
            continue
        author = as_dict(authorship.get("author"))
        raw_name = (author.get("display_name", "") or authorship.get("raw_author_name", "") or "").strip()
        if raw_name:
            names.append(raw_name)
    return "; ".join(names)


def extract_keywords(work: Dict[str, Any]) -> str:
    values: List[str] = []
    for item in work.get("keywords", []):
        if isinstance(item, dict) and item.get("display_name"):
            values.append(str(item["display_name"]).strip())
    if not values:
        for item in work.get("topics", [])[:5]:
            if isinstance(item, dict) and item.get("display_name"):
                values.append(str(item["display_name"]).strip())
    seen: Set[str] = set()
    cleaned: List[str] = []
    for v in values:
        n = normalize_text(v)
        if n and n not in seen:
            seen.add(n)
            cleaned.append(v)
    return "; ".join(cleaned)


def query_openalex(search_term: str, from_date: str) -> List[Dict[str, Any]]:
    works: List[Dict[str, Any]] = []
    cursor    = "*"
    page_count = 0
    headers   = build_headers()

    while True:
        params = {
            "search": f'"{search_term}"',
            "filter": f"from_publication_date:{from_date}",
            "per-page": 200,
            "cursor": cursor,
        }
        try:
            resp = requests.get("https://api.openalex.org/works", params=params, headers=headers, timeout=30)
            resp.raise_for_status()
        except requests.RequestException as e:
            print(f"  Error OpenAlex '{search_term}': {e}")
            break

        data  = resp.json()
        batch = data.get("results", [])
        if not isinstance(batch, list):
            break
        works.extend(batch)

        meta       = as_dict(data.get("meta"))
        next_cursor = meta.get("next_cursor")
        page_count += 1
        if not next_cursor:
            break
        if MAX_PAGES_PER_TERM is not None and page_count >= MAX_PAGES_PER_TERM:
            break
        cursor = next_cursor
        time.sleep(1)

    return works


def extract_record_info(work: Dict[str, Any], search_term: str) -> Dict[str, Any]:
    ids         = as_dict(work.get("ids"))
    openalex_id = str(work.get("id", "")).strip()
    doi         = normalize_doi(ids.get("doi"))
    primary     = as_dict(work.get("primary_location"))
    source      = as_dict(primary.get("source"))
    host_venue  = as_dict(work.get("host_venue"))
    origin      = source.get("display_name") or host_venue.get("display_name") or ""
    doc_type    = primary.get("type") or work.get("type_crossref") or work.get("type") or ""
    title       = (work.get("title") or work.get("display_name") or "").strip()
    oa_info     = as_dict(work.get("open_access"))
    is_oa       = bool(oa_info.get("is_oa", False))
    pdf_url     = primary.get("pdf_url") or oa_info.get("oa_url") or ""
    url         = (primary.get("landing_page_url") or primary.get("pdf_url")
                   or (f"https://doi.org/{doi}" if doi else "") or openalex_id)
    record_id   = doi or openalex_id

    return {
        "record_id":        record_id,
        "first_seen_date":  date.today().isoformat(),
        "search_term":      search_term,
        "source":           "OpenAlex",
        "origin":           origin,
        "document_type":    doc_type,
        "authors":          extract_authors(work),
        "title":            title,
        "abstract":         reconstruct_abstract(work.get("abstract_inverted_index")),
        "keywords":         extract_keywords(work),
        "publication_year": work.get("publication_year") or "",
        "publication_date": work.get("publication_date") or "",
        "doi":              doi,
        "url":              url,
        "openalex_id":      openalex_id,
        "is_oa":            is_oa,
        "pdf_url":          pdf_url,
    }


# ── OAI-PMH (CONICET / SEDICI / RIAA) ────────────────────────────────────────

def _dc_texts(metadata: ET.Element, field: str) -> List[str]:
    els = metadata.findall(f"dc:{field}", OAI_NS)
    return [el.text.strip() for el in els if el.text and el.text.strip()]


def query_oai_pmh(
    base_url:  str,
    from_date: str,
    search_terms: List[str],
    source_name:  str,
    max_pages:    int = 50,
) -> List[Dict[str, Any]]:
    """
    Consulta un repositorio OAI-PMH, descarga registros desde `from_date`
    y filtra localmente por `search_terms` en título, subjects y descripción.
    """
    terms_norm = [normalize_text(t) for t in search_terms]
    records: List[Dict[str, Any]] = []
    token: Optional[str] = None
    page = 0

    print(f"  Consultando {source_name}...")

    while True:
        params = (
            {"verb": "ListRecords", "resumptionToken": token}
            if token
            else {"verb": "ListRecords", "metadataPrefix": "oai_dc", "from": from_date}
        )

        try:
            resp = requests.get(base_url, params=params, timeout=60,
                                headers={"User-Agent": "academic-scraper/3.0"})
            resp.raise_for_status()
        except Exception as e:
            print(f"    Error en {source_name} (pág {page}): {e}")
            break

        try:
            root = ET.fromstring(resp.content)
        except ET.ParseError as e:
            print(f"    XML inválido en {source_name}: {e}")
            break

        # Verificar error OAI
        error_el = root.find(".//oai:error", OAI_NS)
        if error_el is not None:
            print(f"    OAI error {source_name}: {error_el.get('code')} – {error_el.text}")
            break

        for record_el in root.findall(".//oai:record", OAI_NS):
            header = record_el.find("oai:header", OAI_NS)
            if header is not None and header.get("status") == "deleted":
                continue

            metadata = record_el.find(".//oai_dc:dc", OAI_NS)
            if metadata is None:
                continue

            titles       = _dc_texts(metadata, "title")
            if not titles:
                continue
            title = titles[0]

            subjects     = _dc_texts(metadata, "subject")
            descriptions = _dc_texts(metadata, "description")
            dates        = _dc_texts(metadata, "date")
            identifiers  = _dc_texts(metadata, "identifier")
            creators     = _dc_texts(metadata, "creator")
            sources      = _dc_texts(metadata, "source")
            types        = _dc_texts(metadata, "type")

            # Filtro por términos de búsqueda
            searchable = normalize_text(
                " ".join(titles + subjects + descriptions[:1])
            )
            if not any(term in searchable for term in terms_norm):
                continue

            # Año
            year = ""
            for d in dates:
                m = re.search(r"\b(20\d{2})\b", d)
                if m:
                    year = m.group(1)
                    break
            if year and int(year) < 2020:
                continue

            # DOI y URL
            doi = ""
            url = ""
            for ident in identifiers:
                ident_lower = ident.lower()
                if "doi.org" in ident_lower or re.match(r"^10\.\d{4,}/", ident):
                    doi = normalize_doi(ident)
                elif ident.startswith("http") and not url:
                    url = ident
            if not url and doi:
                url = f"https://doi.org/{doi}"

            # OAI identifier como fallback de record_id
            id_el     = record_el.find(".//oai:identifier", OAI_NS)
            oai_id    = (id_el.text or "").strip() if id_el is not None else ""
            record_id = doi or oai_id
            if not record_id:
                continue

            records.append({
                "record_id":        record_id,
                "first_seen_date":  date.today().isoformat(),
                "search_term":      source_name,
                "source":           source_name,
                "origin":           sources[0] if sources else source_name,
                "document_type":    types[0] if types else "",
                "authors":          "; ".join(creators),
                "title":            title,
                "abstract":         descriptions[0] if descriptions else "",
                "keywords":         "; ".join(subjects[:10]),
                "publication_year": year,
                "publication_date": dates[0] if dates else "",
                "doi":              doi,
                "url":              url,
                "openalex_id":      "",
                "is_oa":            True,
                "pdf_url":          "",
            })

        # Paginación OAI
        token_el = root.find(".//oai:resumptionToken", OAI_NS)
        if token_el is not None and token_el.text and token_el.text.strip():
            token  = token_el.text.strip()
            page  += 1
            print(f"    {source_name} pág {page} — {len(records)} registros relevantes...")
            time.sleep(2)
            if page >= max_pages:
                print(f"    {source_name}: límite de páginas alcanzado ({max_pages})")
                break
        else:
            break

    print(f"  {source_name}: {len(records)} registros con términos de búsqueda")
    return records




def query_conicet_rest(from_year: int = 2020) -> List[Dict[str, Any]]:
    """
    Consulta CONICET Digital via su API REST de DSpace.
    Busca por cada término directamente en el motor de búsqueda.
    Mucho más preciso que OAI-PMH con filtro local.
    """
    all_records: List[Dict[str, Any]] = []
    seen_handles: Set[str] = set()

    print("  Consultando CONICET Digital (REST API)...")

    for term in CONICET_SEARCH_TERMS:
        offset = 0
        limit  = 100
        term_count = 0

        while True:
            try:
                resp = requests.get(
                    f"{CONICET_REST_URL}/items",
                    params={
                        "query":  term,
                        "limit":  limit,
                        "offset": offset,
                        "expand": "metadata",
                    },
                    headers={"Accept": "application/json",
                             "User-Agent": "academic-scraper/3.0"},
                    timeout=30,
                )
                resp.raise_for_status()
                items = resp.json()
            except Exception as e:
                print(f"    Error CONICET REST para '{term}': {e}")
                break

            if not items:
                break

            for item in items:
                handle = str(item.get("handle", "") or "").strip()
                if not handle or handle in seen_handles:
                    continue

                # Extraer metadata DC
                meta: Dict[str, List[str]] = {}
                for m in item.get("metadata", []) or []:
                    key = m.get("key", "")
                    val = (m.get("value") or "").strip()
                    if key and val:
                        meta.setdefault(key, []).append(val)

                title = (meta.get("dc.title", [""]) or [""])[0].strip()
                if not title:
                    continue

                # Filtrar por tipo de documento (incluir artículos, tesis y capítulos)
                ACCEPTED_TYPES = {
                    "article", "journal article", "artículo", "articulo",
                    "thesis", "doctoral thesis", "tesis doctoral", "tesis",
                    "book chapter", "capítulo de libro", "book section",
                    "conference paper", "conference object", "ponencia",
                }
                types_list_check = meta.get("dc.type", []) or []
                if types_list_check:
                    type_lower = " ".join(str(t).lower() for t in types_list_check)
                    if not any(t in type_lower for t in ACCEPTED_TYPES):
                        continue

                # Filtrar por año
                dates_raw = meta.get("dc.date.issued", []) or meta.get("dc.date", []) or []
                year = ""
                for d in dates_raw:
                    m_yr = re.search(r"(20\d{2})", str(d))
                    if m_yr:
                        year = m_yr.group(1)
                        break
                if year and int(year) < from_year:
                    continue

                # DOI y URL
                doi = ""
                identifiers = meta.get("dc.identifier.uri", []) + meta.get("dc.identifier", [])
                url = f"https://ri.conicet.gov.ar/handle/{handle}"
                for ident in identifiers:
                    ident_lower = str(ident).lower()
                    if "doi.org" in ident_lower or re.match(r"^10\.\d{{4,}}/", str(ident)):
                        doi = normalize_doi(str(ident))
                    elif str(ident).startswith("http") and "ri.conicet" in str(ident).lower():
                        url = str(ident)

                record_id = doi or f"conicet:{handle}"
                seen_handles.add(handle)

                authors_list = meta.get("dc.contributor.author", []) or meta.get("dc.creator", []) or []
                subjects_list = (meta.get("dc.subject", []) or [])[:10]
                abstract_list = meta.get("dc.description.abstract", []) or meta.get("dc.description", []) or []
                types_list    = meta.get("dc.type", []) or []
                source_list   = meta.get("dc.relation.journal", []) or meta.get("dc.publisher", []) or []

                all_records.append({
                    "record_id":        record_id,
                    "first_seen_date":  date.today().isoformat(),
                    "search_term":      term,
                    "source":           "CONICET Digital",
                    "origin":           source_list[0] if source_list else "CONICET Digital",
                    "document_type":    types_list[0] if types_list else "",
                    "authors":          "; ".join(str(a) for a in authors_list),
                    "title":            title,
                    "abstract":         abstract_list[0] if abstract_list else "",
                    "keywords":         "; ".join(str(s) for s in subjects_list),
                    "publication_year": year,
                    "publication_date": dates_raw[0] if dates_raw else "",
                    "doi":              doi,
                    "url":              url,
                    "openalex_id":      "",
                    "is_oa":            True,
                    "pdf_url":          "",
                })
                term_count += 1

            if len(items) < limit:
                break
            offset += limit
            time.sleep(1)

        print(f"    CONICET '{term}': {term_count} registros")

    print(f"  CONICET Digital total: {len(all_records)} registros")
    return all_records

def fetch_oa_info_by_doi(doi: str) -> Tuple[bool, str]:
    if not doi:
        return False, ""
    try:
        resp = requests.get(
            f"https://api.openalex.org/works/https://doi.org/{doi}",
            headers=build_headers(), timeout=15,
        )
        if resp.status_code != 200:
            return False, ""
        data    = resp.json()
        oa      = as_dict(data.get("open_access"))
        primary = as_dict(data.get("primary_location"))
        pdf_url = primary.get("pdf_url") or oa.get("oa_url") or ""
        return bool(oa.get("is_oa", False)), str(pdf_url)
    except Exception:
        return False, ""


# ── Recolección principal ─────────────────────────────────────────────────────

def collect_new_records() -> Tuple[List[Dict[str, Any]], int]:
    ensure_storage()

    seen_ids             = load_seen_ids()
    title_year_sigs      = load_existing_signatures()
    fuzzy_index          = build_fuzzy_index()
    new_records: List[Dict[str, Any]] = []
    # Índice en memoria de lo que ya agregamos en esta corrida (para dedup entre fuentes)
    session_index: Dict[str, List[Tuple[str, str]]] = {}

    def _is_new(record: Dict[str, Any]) -> bool:
        """Aplica las tres capas de dedup."""
        rid  = str(record.get("record_id", "")).strip()
        sig  = build_signature(record.get("title", ""), record.get("publication_year", ""))
        year = str(record.get("publication_year", "") or "").strip()
        tnorm = normalize_text(record.get("title", ""))
        fauth = get_first_author(record.get("authors", ""))

        if not rid:
            return False
        # Capa 1: DOI/ID exacto
        if rid in seen_ids:
            return False
        # Capa 2a: firma exacta (título normalizado + año)
        if sig in title_year_sigs:
            seen_ids.add(rid)
            return False
        # Capa 2b: fuzzy contra base existente
        if is_fuzzy_duplicate(record, fuzzy_index):
            seen_ids.add(rid)
            return False
        # Capa 2c: fuzzy contra registros de esta misma corrida
        if is_fuzzy_duplicate(record, session_index):
            seen_ids.add(rid)
            return False
        return True

    def _register(record: Dict[str, Any]) -> None:
        rid  = str(record.get("record_id", ""))
        sig  = build_signature(record.get("title", ""), record.get("publication_year", ""))
        year = str(record.get("publication_year", "") or "").strip()
        seen_ids.add(rid)
        title_year_sigs.add(sig)
        tnorm = normalize_text(record.get("title", ""))
        fauth = get_first_author(record.get("authors", ""))
        session_index.setdefault(year, []).append((tnorm, fauth))

    # ── 1. OpenAlex ──────────────────────────────────────────────────────────
    print("\n=== OpenAlex ===")
    for term in SEARCH_TERMS:
        print(f"Buscando: {term}")
        works = query_openalex(term, START_DATE)
        print(f"  {len(works)} resultados")
        for work in works:
            if not isinstance(work, dict):
                continue
            rec = extract_record_info(work, term)
            if _is_new(rec):
                new_records.append(rec)
                _register(rec)

    # ── 2. Repositorios OAI-PMH ───────────────────────────────────────────────
    print("\n=== Repositorios institucionales (OAI-PMH) ===")
    for src in OAI_SOURCES:
        oai_records = query_oai_pmh(
            base_url     = src["url"],
            from_date    = START_DATE,
            search_terms = SEARCH_TERMS,
            source_name  = src["name"],
            max_pages    = src["max_pages"],
        )
        for rec in oai_records:
            if _is_new(rec):
                new_records.append(rec)
                _register(rec)

    append_master_records(new_records)
    save_seen_ids(seen_ids)

    total = count_master_records()
    print(f"\nNuevos registros: {len(new_records)} | Total acumulado: {total}")
    return new_records, total


# ── Excel ─────────────────────────────────────────────────────────────────────

EXCEL_HEADERS = [
    "Título", "Año", "Fecha publicación", "Autores", "Revista/Fuente",
    "Fuente", "Tipo", "Término búsqueda", "DOI", "URL", "Acceso abierto",
    "PDF URL", "Abstract", "Palabras clave", "Fecha registro",
]
EXCEL_KEYS = [
    "title", "publication_year", "publication_date", "authors", "origin",
    "source", "document_type", "search_term", "doi", "url", "is_oa",
    "pdf_url", "abstract", "keywords", "first_seen_date",
]


def update_excel(new_record_ids: Set[str]) -> None:
    all_records = read_all_records()
    if not all_records:
        print("ADVERTENCIA: No se encontraron registros para generar el Excel.")
        return
    print(f"Generando Excel con {len(all_records)} registros...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Publicaciones"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    new_fill = PatternFill("solid", fgColor="FFF2CC")
    for row_idx, record in enumerate(all_records, 2):
        is_new = record.get("record_id", "") in new_record_ids
        for col_idx, key in enumerate(EXCEL_KEYS, 1):
            val = record.get(key, "") or ""
            if key == "is_oa":
                val = "Sí" if str(val).lower() in ("true", "1", "yes", "sí") else "No"
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=(key in ("title", "abstract", "keywords")))
            if is_new:
                cell.fill = new_fill
    col_widths = [55, 6, 14, 40, 30, 14, 12, 18, 38, 38, 12, 38, 60, 40, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    try:
        wb.save(EXCEL_FILE)
        print(f"Excel guardado: {len(all_records)} registros")
    except Exception as e:
        print(f"Error al guardar Excel: {e}")


# ── PDFs ──────────────────────────────────────────────────────────────────────

def download_pdf(record: Dict[str, Any]) -> Optional[Path]:
    pdf_url = str(record.get("pdf_url", "")).strip()
    if not pdf_url:
        return None
    raw_name = (
        str(record.get("doi") or record.get("record_id") or "unknown")
        .replace("/", "_").replace(":", "_").replace(" ", "_")
    )
    dest = PDFS_DIR / f"{raw_name}.pdf"
    if dest.exists():
        return dest
    try:
        resp = requests.get(pdf_url, timeout=30, stream=True)
        resp.raise_for_status()
        content_type = resp.headers.get("content-type", "")
        if "pdf" not in content_type and not pdf_url.lower().endswith(".pdf"):
            return None
        with open(dest, "wb") as f:
            for chunk in resp.iter_content(8192):
                f.write(chunk)
        print(f"  PDF descargado: {dest.name}")
        return dest
    except Exception as exc:
        print(f"  No se pudo descargar ({pdf_url}): {exc}")
        return None


def bulk_download_and_upload(uploader: Optional[Any]) -> None:
    all_records = read_all_records()
    print(f"\nDescarga masiva de PDFs: {len(all_records)} registros...")
    ok = skipped = 0
    for i, record in enumerate(all_records):
        pdf_url  = str(record.get("pdf_url", "")).strip()
        is_oa    = str(record.get("is_oa", "")).lower() in ("true", "1", "yes", "sí")
        if not pdf_url and not is_oa:
            doi = str(record.get("doi", "")).strip()
            if doi:
                is_oa, pdf_url = fetch_oa_info_by_doi(doi)
                time.sleep(0.3)
        if not pdf_url:
            skipped += 1
            continue
        pdf_path = download_pdf({**record, "pdf_url": pdf_url})
        if pdf_path:
            ok += 1
            if uploader:
                try:
                    uploader.upload(pdf_path)
                except Exception as exc:
                    print(f"  Error subiendo a Drive: {exc}")
        if (i + 1) % 100 == 0:
            print(f"  {i+1}/{len(all_records)} — PDFs: {ok}")
    print(f"PDFs descargados: {ok} | Sin PDF: {skipped}")


# ── Correo ────────────────────────────────────────────────────────────────────

def generate_email_body(new_records: List[Dict[str, Any]], total_records: int) -> str:
    lines = [
        f"Informe diario – dirección/gestión escolar – {date.today().isoformat()}",
        "",
        f"Base acumulada: {total_records} registros",
        f"Nuevos en esta corrida: {len(new_records)}",
        "",
    ]
    if not new_records:
        lines += ["Sin novedades. Excel adjunto actualizado."]
        return "\n".join(lines)

    # Agrupar por fuente
    from collections import Counter
    fuentes = Counter(r.get("source", "OpenAlex") for r in new_records)
    lines.append("Fuentes:")
    for fuente, cnt in fuentes.most_common():
        lines.append(f"  {fuente}: {cnt} nuevos")
    lines.append("")

    sorted_records = sorted(
        new_records,
        key=lambda x: (str(x.get("publication_date") or ""), int(x.get("publication_year") or 0)),
        reverse=True,
    )
    lines.append(f"Hasta {EMAIL_ITEM_LIMIT} novedades (ver Excel para el resto):")
    lines.append("")
    for i, rec in enumerate(sorted_records[:EMAIL_ITEM_LIMIT], 1):
        lines += [
            f"{i}. {rec.get('title','Sin título')} ({rec.get('publication_year','s/f')})",
            f"   Autores: {rec.get('authors','—')}",
            f"   Fuente: {rec.get('source','—')} | {rec.get('origin','—')}",
            f"   DOI: {rec.get('doi','Sin DOI')}",
            f"   URL: {rec.get('url','Sin URL')}",
            "",
        ]
    if len(sorted_records) > EMAIL_ITEM_LIMIT:
        lines.append(f"... y {len(sorted_records) - EMAIL_ITEM_LIMIT} más en el Excel adjunto.")
    return "\n".join(lines)


def send_email(subject: str, body: str, gmail_user: str, gmail_password: str, recipient: str) -> None:
    msg = MIMEMultipart()
    msg["Subject"] = subject
    msg["From"]    = gmail_user
    msg["To"]      = recipient
    msg.attach(MIMEText(body, _subtype="plain", _charset="utf-8"))
    if EXCEL_FILE.exists():
        with open(EXCEL_FILE, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={EXCEL_FILE.name}")
        msg.attach(part)
    with smtplib.SMTP("smtp.gmail.com", 587, timeout=30) as server:
        server.ehlo(); server.starttls(); server.ehlo()
        server.login(gmail_user, gmail_password)
        server.sendmail(gmail_user, [recipient], msg.as_string())


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    new_records, total_records = collect_new_records()
    new_record_ids = {str(r.get("record_id", "")) for r in new_records}

    update_excel(new_record_ids)

    drive_folder_id = os.getenv("DRIVE_FOLDER_ID", "").strip()
    uploader = None
    if DRIVE_AVAILABLE and drive_folder_id:
        try:
            uploader = DriveUploader(drive_folder_id)
            print("Google Drive conectado.")
        except Exception as exc:
            print(f"Drive no disponible: {exc}")

    bulk_download_and_upload(uploader)

    gmail_user     = os.getenv("GMAIL_USER", "").strip()
    gmail_password = (os.getenv("GMAIL_APP_PASSWORD", "") or os.getenv("GMAIL_PASSWORD", "")).strip()
    recipient      = os.getenv("RECIPIENT_EMAIL", "").strip()
    body           = generate_email_body(new_records, total_records)
    subject        = "Informe diario – dirección/gestión escolar"

    if gmail_user and gmail_password and recipient:
        send_email(subject, body, gmail_user, gmail_password, recipient)
        print("Correo enviado.")
    else:
        print("Faltan credenciales de correo. Informe en consola:")
        print(body)


if __name__ == "__main__":
    main()
