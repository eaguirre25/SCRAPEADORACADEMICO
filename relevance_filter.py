#!/usr/bin/env python3
"""
Filtro de pertinencia para el Scrapeador Académico.

Objetivo:
- Evitar que resultados amplios de OpenAlex, CONICET Digital, SEDICI o RIAA
  entren al dashboard si no tienen relación clara con dirección escolar.
- Mantener un archivo de revisión para casos dudosos.
- Mantener un archivo de rechazados para auditar por qué algo quedó afuera.

El filtro trabaja después de main.py y antes de generar el dashboard.
"""
from __future__ import annotations

import csv
import json
import re
import unicodedata
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DATA_DIR = Path("data")
MASTER_CSV = DATA_DIR / "master_records.csv"
REVIEW_CSV = DATA_DIR / "review_records.csv"
REJECTED_CSV = DATA_DIR / "rejected_records.csv"
LATEST_RELEVANT_CSV = DATA_DIR / "latest_relevant_records.csv"
AUTO_REVIEW_PROMOTED_CSV = DATA_DIR / "auto_review_promoted.csv"
AUTO_REVIEW_REJECTED_CSV = DATA_DIR / "auto_review_rejected.csv"
REPORT_JSON = DATA_DIR / "relevance_filter_report.json"
EXCEL_FILE = DATA_DIR / "publicaciones.xlsx"

CSV_FIELDS = [
    "record_id", "first_seen_date", "search_term", "source",
    "origin", "document_type", "authors", "title", "abstract",
    "keywords", "publication_year", "publication_date",
    "doi", "url", "openalex_id", "is_oa", "pdf_url",
]

# Frases que expresan relación directa con el objeto de interés.
STRONG_PHRASES = [
    "direccion escolar", "dirección escolar",
    "gestion escolar", "gestión escolar",
    "liderazgo escolar",
    "equipo directivo", "equipos directivos",
    "directivos escolares", "directivo escolar", "directiva escolar",
    "director escolar", "directora escolar", "director de escuela", "directora de escuela",
    "conduccion escolar", "conducción escolar",
    "gobierno escolar", "gobierno de la escuela",
    "school leadership", "school management", "school governance",
    "school principal", "school principals", "principalship", "headteacher", "head teacher",
    "school administration",
]

# Frases pertinentes pero demasiado amplias si aparecen solas.
MEDIUM_PHRASES = [
    "gestion educativa", "gestión educativa",
    "liderazgo educativo", "educational leadership",
    "administracion escolar", "administración escolar",
    "educational management",
]

ROLE_TERMS = [
    "director", "directora", "directores", "directoras",
    "directivo", "directiva", "directivos", "directivas",
    "equipo directivo", "equipos directivos",
    "principal", "principals", "headteacher", "head teacher",
]

MANAGEMENT_TERMS = [
    "gestion", "gestión", "management",
    "liderazgo", "leadership",
    "conduccion", "conducción",
    "gobierno", "governance",
    "administracion", "administración", "administration",
]

CORE_SCHOOL_TERMS = [
    "escuela", "escuelas", "escolar", "escolares",
    "secundaria", "secundario", "secondary school", "secondary education",
    "school", "schools", "schooling",
]

EDUCATION_CONTEXT_TERMS = [
    "educacion", "educación", "educativa", "educativo",
    "institucion educativa", "institución educativa",
    "instituciones educativas",
    "docente", "docentes", "profesor", "profesores",
    "estudiante", "estudiantes", "alumno", "alumnos",
]

SCHOOL_SYSTEM_TERMS = [
    "educacion basica", "educación básica", "basic education",
    "educacion secundaria", "educación secundaria", "secondary education",
    "educacion primaria", "educación primaria", "primary education",
    "bachillerato", "k-12", "k12",
    "colegio", "colegios", "liceo", "liceos",
    "centro educativo", "centros educativos",
    "institucion educativa", "institución educativa", "instituciones educativas",
    "unidad educativa", "unidades educativas",
    "comunidad educativa", "comunidades educativas",
]

SCHOOL_AUTHORITY_TERMS = [
    "autoridad", "autoridades",
    "administrador escolar", "administradores escolares",
    "gestion institucional", "gestión institucional",
    "gestion administrativa", "gestión administrativa",
    "school administrator", "school administrators",
    "school leader", "school leaders",
    "management committee", "management committees",
]

HIGHER_ED_NOISE_TERMS = [
    "universidad", "universitario", "universitaria", "universitarias",
    "higher education", "educacion superior", "educación superior",
    "postgrado", "postgrados", "posgrado", "posgrados",
]

SECTOR_NOISE_TERMS = [
    "educacion fisica", "educación física", "physical education",
    "salud", "health", "medicine", "medical",
    "sexuality education", "educacion sexual", "educación sexual",
    "covid", "pandemia", "pandemic",
]

OWNERSHIP_MANAGEMENT_NOISE_TERMS = [
    "gestion privada", "gestión privada",
    "gestion publica", "gestión pública",
    "gestion estatal", "gestión estatal",
    "gestion social", "gestión social",
    "public management", "private management",
]

# Ruido frecuente: solo penaliza cuando no hay señales fuertes de dirección escolar.
NOISE_TERMS = [
    "tourism", "turismo", "penguin", "penguins", "pinguino", "pingüino",
    "magellanic", "conservation", "conservacion", "conservación",
    "competencias parentales", "parentales percibidas", "parental competencies",
    "salud", "health", "agricultura", "agriculture",
]

STRICT_SOURCES = {"CONICET Digital", "SEDICI-UNLP", "RIAA-UNSAM", "Repositorio / Otro"}


def strip_accents(text: str) -> str:
    text = unicodedata.normalize("NFD", text)
    return "".join(ch for ch in text if unicodedata.category(ch) != "Mn")


def norm(value: Any) -> str:
    text = str(value or "").casefold()
    text = strip_accents(text)
    text = re.sub(r"[^\w\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def phrase_hit(text: str, phrases: Iterable[str]) -> List[str]:
    hits: List[str] = []
    padded = f" {text} "
    for phrase in phrases:
        p = norm(phrase)
        if p and f" {p} " in padded:
            hits.append(phrase)
    return hits


def term_hit(text: str, terms: Iterable[str]) -> List[str]:
    hits: List[str] = []
    padded = f" {text} "
    for term in terms:
        t = norm(term)
        if not t:
            continue
        if " " in t:
            ok = f" {t} " in padded
        else:
            ok = re.search(rf"\b{re.escape(t)}\b", text) is not None
        if ok:
            hits.append(term)
    return hits


def row_text(row: Dict[str, Any]) -> Tuple[str, str, str]:
    title = norm(row.get("title"))
    body = norm(" ".join([
        str(row.get("abstract") or ""),
        str(row.get("keywords") or ""),
        str(row.get("document_type") or ""),
    ]))
    all_text = f"{title} {body}".strip()
    return title, body, all_text


def source_name(row: Dict[str, Any]) -> str:
    return str(row.get("source") or "Repositorio / Otro").strip() or "Repositorio / Otro"


def source_is_strict(source: str) -> bool:
    parts = [part.strip() for part in source.split("|")]
    return any(part in STRICT_SOURCES for part in parts)


def classify_relevance(row: Dict[str, Any]) -> Tuple[str, int, str, List[str]]:
    """Devuelve categoría, puntaje, motivo y evidencias."""
    title, body, text = row_text(row)
    source = source_name(row)

    strong_title = phrase_hit(title, STRONG_PHRASES)
    strong_body = phrase_hit(body, STRONG_PHRASES)
    medium_title = phrase_hit(title, MEDIUM_PHRASES)
    medium_body = phrase_hit(body, MEDIUM_PHRASES)

    role_hits = term_hit(text, ROLE_TERMS)
    management_hits = term_hit(text, MANAGEMENT_TERMS)
    core_school_hits = term_hit(text, CORE_SCHOOL_TERMS)
    education_hits = term_hit(text, EDUCATION_CONTEXT_TERMS)
    higher_ed_noise = term_hit(text, HIGHER_ED_NOISE_TERMS)
    sector_noise = term_hit(text, SECTOR_NOISE_TERMS)
    ownership_noise = term_hit(text, OWNERSHIP_MANAGEMENT_NOISE_TERMS)
    noise_hits = term_hit(text, NOISE_TERMS)

    evidence: List[str] = []
    score = 0

    if strong_title:
        score += 7
        evidence.append("frase fuerte en título: " + ", ".join(strong_title[:3]))
    if strong_body:
        score += 5
        evidence.append("frase fuerte en resumen/palabras clave: " + ", ".join(strong_body[:3]))
    if medium_title:
        score += 3
        evidence.append("frase media en título: " + ", ".join(medium_title[:3]))
    if medium_body:
        score += 2
        evidence.append("frase media en resumen/palabras clave: " + ", ".join(medium_body[:3]))

    if role_hits and (core_school_hits or education_hits):
        score += 5
        evidence.append("rol directivo + contexto educativo/escolar")
    if management_hits and core_school_hits:
        score += 4
        evidence.append("gestión/liderazgo/gobierno + escuela/escolaridad")
    if management_hits and education_hits and not core_school_hits:
        score += 1
        evidence.append("gestión/liderazgo + educación genérica")

    if noise_hits and not (strong_title or strong_body or (role_hits and (core_school_hits or education_hits))):
        score -= 4
        evidence.append("ruido temático: " + ", ".join(noise_hits[:3]))

    strict = source_is_strict(source)
    direct_focus = bool(strong_title or strong_body or (role_hits and (core_school_hits or education_hits)))

    if ownership_noise and not direct_focus:
        score -= 3
        evidence.append("gestión como dependencia/sector, no como objeto directivo: " + ", ".join(ownership_noise[:3]))

    if (higher_ed_noise or sector_noise) and not direct_focus:
        evidence.append("ruido de nivel/sector: " + ", ".join((higher_ed_noise + sector_noise)[:3]))
        return "rechazada", score, "fuera de foco: nivel/sector sin dirección o gestión escolar como objeto", evidence

    # Alta pertinencia: lo que entra al master y al dashboard.
    if strong_title or strong_body:
        return "alta", score, "coincidencia directa con dirección/gestión/liderazgo escolar", evidence
    if role_hits and (core_school_hits or education_hits):
        return "alta", score, "menciona rol/equipo directivo en contexto educativo", evidence
    if strict and management_hits and core_school_hits:
        return "revisar", score, "fuente externa con gestión + escuela, requiere confirmar foco directivo", evidence
    if management_hits and core_school_hits and score >= 4:
        return "alta", score, "articula gestión/liderazgo/gobierno con escuela/escolaridad", evidence

    # En fuentes externas a OpenAlex, las coincidencias amplias quedan en revisión.
    if strict:
        if score >= 3:
            return "revisar", score, "fuente externa con coincidencia amplia, requiere revisión", evidence
        return "rechazada", score, "sin señales suficientes de dirección escolar", evidence

    # OpenAlex puede traer resúmenes más ricos, pero gestión educativa sola no alcanza.
    if medium_title or medium_body:
        return "revisar", score, "coincidencia amplia con gestión/liderazgo educativo", evidence
    if score >= 3:
        return "revisar", score, "coincidencia parcial, requiere revisión", evidence

    return "rechazada", score, "sin relación clara con dirección escolar", evidence


def second_review(row: Dict[str, Any]) -> Tuple[str, str, List[str]]:
    """Resuelve casos en revisión con reglas conservadoras de segunda lectura."""
    title, body, text = row_text(row)
    search_term = norm(row.get("search_term"))
    evidence: List[str] = []

    strong_hits = phrase_hit(text, STRONG_PHRASES)
    medium_hits = phrase_hit(text, MEDIUM_PHRASES)
    role_hits = term_hit(text, ROLE_TERMS)
    management_hits = term_hit(text, MANAGEMENT_TERMS)
    core_school_hits = term_hit(text, CORE_SCHOOL_TERMS)
    school_system_hits = term_hit(text, SCHOOL_SYSTEM_TERMS)
    school_authority_hits = term_hit(text, SCHOOL_AUTHORITY_TERMS)
    higher_ed_noise = term_hit(text, HIGHER_ED_NOISE_TERMS)
    sector_noise = term_hit(text, SECTOR_NOISE_TERMS)
    noise_hits = term_hit(text, NOISE_TERMS)

    if strong_hits:
        evidence.append("segunda revisión: frase directa: " + ", ".join(strong_hits[:3]))
        return "promover", "segunda revisión confirma relación directa", evidence

    if role_hits and (core_school_hits or school_system_hits):
        evidence.append("segunda revisión: rol directivo + sistema escolar")
        return "promover", "segunda revisión confirma rol directivo en contexto escolar", evidence

    if school_authority_hits and (core_school_hits or school_system_hits) and management_hits:
        evidence.append("segunda revisión: autoridad/gestión institucional + escuela")
        return "promover", "segunda revisión confirma gestión institucional escolar", evidence

    if management_hits and core_school_hits and not (higher_ed_noise or sector_noise):
        evidence.append("segunda revisión: gestión/liderazgo + escuela sin ruido sectorial")
        return "promover", "segunda revisión confirma gestión escolar", evidence

    broad_only = bool(medium_hits) and not (role_hits or core_school_hits or school_system_hits or school_authority_hits)
    if broad_only:
        evidence.append("segunda revisión: coincidencia demasiado amplia: " + ", ".join(medium_hits[:3]))
        return "descartar", "gestión/liderazgo educativo sin foco escolar/directivo", evidence

    if higher_ed_noise and not (core_school_hits or school_system_hits):
        evidence.append("segunda revisión: foco en educación superior/postgrado")
        return "descartar", "fuera de dirección escolar: educación superior o postgrado", evidence

    if sector_noise and not (role_hits or school_authority_hits or strong_hits):
        evidence.append("segunda revisión: foco sectorial no directivo: " + ", ".join(sector_noise[:3]))
        return "descartar", "fuera de dirección escolar: tema educativo sectorial", evidence

    if noise_hits and not (role_hits or school_authority_hits or strong_hits):
        evidence.append("segunda revisión: ruido temático: " + ", ".join(noise_hits[:3]))
        return "descartar", "fuera de dirección escolar: ruido temático", evidence

    if search_term in {norm(term) for term in MEDIUM_PHRASES} and not (core_school_hits or school_system_hits):
        evidence.append("segunda revisión: término de búsqueda amplio sin anclaje escolar")
        return "descartar", "coincidencia por término amplio, sin anclaje escolar suficiente", evidence

    evidence.append("segunda revisión: requiere lectura humana")
    return "duda", "caso ambiguo para revisión manual", evidence


def read_csv_if_exists(path: Path) -> Tuple[List[str], List[Dict[str, str]]]:
    if not path.exists():
        return [], []
    csv.field_size_limit(20_000_000)
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        return list(reader.fieldnames or []), list(reader)


def read_master() -> Tuple[List[str], List[Dict[str, str]]]:
    master_fields, master_rows = read_csv_if_exists(MASTER_CSV)
    review_fields, review_rows = read_csv_if_exists(REVIEW_CSV)

    fieldnames = list(master_fields or CSV_FIELDS)
    for field in review_fields:
        if field not in fieldnames:
            fieldnames.append(field)
    for field in CSV_FIELDS:
        if field not in fieldnames:
            fieldnames.append(field)

    rows: List[Dict[str, str]] = []
    seen: set[str] = set()
    for source_rows in [master_rows, review_rows]:
        for row in source_rows:
            rid = str(row.get("record_id") or "").strip()
            fallback = "::".join([
                norm(row.get("title")),
                str(row.get("publication_year") or "").strip(),
                norm(row.get("doi") or row.get("url")),
            ])
            key = rid or fallback
            if key in seen:
                continue
            seen.add(key)
            rows.append(row)
    return fieldnames, rows


def record_key(row: Dict[str, Any]) -> str:
    rid = str(row.get("record_id") or "").strip()
    if rid:
        return rid
    return "::".join([
        norm(row.get("title")),
        str(row.get("publication_year") or "").strip(),
        norm(row.get("doi") or row.get("url")),
    ])


def append_unique(target: List[Dict[str, Any]], rows: List[Dict[str, Any]]) -> None:
    seen = {record_key(row) for row in target}
    for row in rows:
        key = record_key(row)
        if key in seen:
            continue
        seen.add(key)
        target.append(row)


def write_csv(path: Path, fieldnames: List[str], rows: List[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    extra_fields = [
        "relevance_status", "relevance_score", "relevance_reason", "relevance_evidence",
        "second_review_status", "second_review_reason", "second_review_evidence",
    ]
    final_fields = list(fieldnames)
    for field in extra_fields:
        if rows and field in rows[0] and field not in final_fields:
            final_fields.append(field)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=final_fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_excel(relevant: List[Dict[str, Any]], review: List[Dict[str, Any]], rejected: List[Dict[str, Any]], fieldnames: List[str]) -> None:
    wb = openpyxl.Workbook()
    sheets = [
        ("Alta pertinencia", relevant),
        ("Revisar", review),
        ("Rechazados", rejected[:5000]),
    ]
    for idx, (name, rows) in enumerate(sheets):
        ws = wb.active if idx == 0 else wb.create_sheet(title=name)
        ws.title = name
        sheet_fields = list(fieldnames)
        for extra in [
            "relevance_status", "relevance_score", "relevance_reason", "relevance_evidence",
            "second_review_status", "second_review_reason", "second_review_evidence",
        ]:
            if rows and extra in rows[0] and extra not in sheet_fields:
                sheet_fields.append(extra)
        ws.append(sheet_fields)
        for row in rows:
            ws.append([row.get(field, "") for field in sheet_fields])
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        for col_idx, field in enumerate(sheet_fields, start=1):
            width = 18
            if field in {"title", "abstract", "keywords", "relevance_evidence", "second_review_evidence"}:
                width = 55
            elif field in {"url", "pdf_url", "doi"}:
                width = 34
            elif field in {"authors", "origin"}:
                width = 36
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_FILE)


def main() -> None:
    fieldnames, rows = read_master()
    rejected_fields, previously_rejected = read_csv_if_exists(REJECTED_CSV)
    for field in rejected_fields:
        if field not in fieldnames:
            fieldnames.append(field)

    relevant: List[Dict[str, Any]] = []
    review: List[Dict[str, Any]] = []
    rejected: List[Dict[str, Any]] = []
    auto_promoted: List[Dict[str, Any]] = []
    auto_rejected: List[Dict[str, Any]] = []

    today = date.today().isoformat()

    for row in rows:
        status, score, reason, evidence = classify_relevance(row)
        row_out: Dict[str, Any] = dict(row)
        row_out["relevance_status"] = status
        row_out["relevance_score"] = score
        row_out["relevance_reason"] = reason
        row_out["relevance_evidence"] = " | ".join(evidence)
        if status == "alta":
            relevant.append(row_out)
        elif status == "revisar":
            second_status, second_reason, second_evidence = second_review(row_out)
            row_out["second_review_status"] = second_status
            row_out["second_review_reason"] = second_reason
            row_out["second_review_evidence"] = " | ".join(second_evidence)
            if second_status == "promover":
                row_out["relevance_status"] = "alta"
                row_out["relevance_reason"] = f"{reason}; {second_reason}"
                relevant.append(row_out)
                auto_promoted.append(row_out)
            elif second_status == "descartar":
                row_out["relevance_status"] = "rechazada"
                row_out["relevance_reason"] = f"{reason}; {second_reason}"
                rejected.append(row_out)
                auto_rejected.append(row_out)
            else:
                review.append(row_out)
        else:
            rejected.append(row_out)

    append_unique(rejected, previously_rejected)

    latest_relevant = [r for r in relevant if str(r.get("first_seen_date") or "") == today]

    # El master queda curado: solo alta pertinencia.
    write_csv(MASTER_CSV, fieldnames, relevant)
    write_csv(REVIEW_CSV, fieldnames, review)
    write_csv(REJECTED_CSV, fieldnames, rejected)
    write_csv(LATEST_RELEVANT_CSV, fieldnames, latest_relevant)
    write_csv(AUTO_REVIEW_PROMOTED_CSV, fieldnames, auto_promoted)
    write_csv(AUTO_REVIEW_REJECTED_CSV, fieldnames, auto_rejected)
    write_excel(relevant, review, rejected, fieldnames)

    report = {
        "generated": today,
        "input_total": len(rows),
        "kept_high_relevance": len(relevant),
        "review_records": len(review),
        "rejected_records": len(rejected),
        "latest_high_relevance": len(latest_relevant),
        "second_review": {
            "input_review_candidates": len(review) + len(auto_promoted) + len(auto_rejected),
            "auto_promoted_to_high_relevance": len(auto_promoted),
            "auto_rejected": len(auto_rejected),
            "remaining_manual_review": len(review),
        },
        "by_source_input": Counter(source_name(r) for r in rows),
        "by_source_kept": Counter(source_name(r) for r in relevant),
        "by_source_review": Counter(source_name(r) for r in review),
        "by_source_rejected": Counter(source_name(r) for r in rejected),
        "by_source_second_review_promoted": Counter(source_name(r) for r in auto_promoted),
        "by_source_second_review_rejected": Counter(source_name(r) for r in auto_rejected),
        "policy": {
            "master_records_csv": "solo alta pertinencia",
            "review_records_csv": "dudas reales que sobrevivieron a la segunda revisión automática",
            "rejected_records_csv": "descartes automáticos auditables",
            "auto_review_promoted_csv": "casos promovidos automáticamente desde revisión a alta pertinencia",
            "auto_review_rejected_csv": "casos descartados automáticamente en segunda revisión",
        },
    }
    # Convertir Counter a dict normal para JSON.
    for key in [
        "by_source_input", "by_source_kept", "by_source_review", "by_source_rejected",
        "by_source_second_review_promoted", "by_source_second_review_rejected",
    ]:
        report[key] = dict(report[key])

    REPORT_JSON.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print("=== Relevance filter ===")
    print(f"Entrada total: {len(rows)}")
    print(f"Alta pertinencia: {len(relevant)}")
    print(f"Revisar: {len(review)}")
    print(f"Rechazados: {len(rejected)}")
    print(f"Segunda revisión - promovidos: {len(auto_promoted)}")
    print(f"Segunda revisión - descartados: {len(auto_rejected)}")
    print(f"Segunda revisión - dudas reales: {len(review)}")
    print(f"Alta pertinencia de hoy: {len(latest_relevant)}")
    print(f"Reporte: {REPORT_JSON}")


if __name__ == "__main__":
    main()
